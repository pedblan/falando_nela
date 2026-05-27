from __future__ import annotations

import argparse
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from processamento.normalizacao import DATASET_NAME, DATASET_VERSION
from processamento.parquet_xlsx import (
    EXCEL_MAX_CELL_CHARS,
    export_parquets_to_xlsx,
    normalize_excel_value,
    resolve_xlsx_paths,
)


def test_resolve_xlsx_paths_defaults_to_local_sample_roots() -> None:
    args = _args(profile="samples-local")

    parquet_root, output_root = resolve_xlsx_paths(args)

    assert parquet_root == Path(f"data/samples/{DATASET_NAME}/{DATASET_VERSION}/parquet")
    assert output_root == Path(f"data/samples/{DATASET_NAME}/{DATASET_VERSION}/xlsx")


def test_export_parquets_to_xlsx_writes_files_and_manifest(tmp_path: Path) -> None:
    parquet_root = tmp_path / "samples" / "textos_parlamentares" / "v1" / "parquet"
    output_root = tmp_path / "samples" / "textos_parlamentares" / "v1" / "xlsx"
    _write_parquet(
        parquet_root / "camara__ccjc_eventos.parquet",
        {
            "texto_id": ["texto-1", "texto-2"],
            "texto": ["=nao_formula", "Texto" + "\x01" + ("x" * 60)],
            "texto_tamanho": [11, 64],
            "raw_path": ["raw/a.jsonl", "raw/b.jsonl"],
        },
    )

    manifest = export_parquets_to_xlsx(
        parquet_root=parquet_root,
        output_root=output_root,
        max_cell_chars=50,
        overwrite=True,
    )

    xlsx_path = output_root / "camara__ccjc_eventos.xlsx"
    assert xlsx_path.exists()
    assert manifest["output_file_count"] == 1
    assert manifest["rows_written"] == 2
    assert manifest["truncated_cells"] == 1
    assert manifest["sanitized_cells"] == 1
    assert manifest["escaped_formula_like_cells"] == 1
    assert (output_root / "manifest.json").exists()

    workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
    worksheet = workbook["camara__ccjc_eventos"]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("texto_id", "texto", "texto_tamanho", "raw_path")
    assert rows[1][1] == "'=nao_formula"
    assert rows[2][1].endswith(" [TRUNCADO PARA XLSX]")
    assert len(rows[2][1]) == 50
    assert "\x01" not in rows[2][1]


def test_export_parquets_to_xlsx_can_drop_columns(tmp_path: Path) -> None:
    parquet_root = tmp_path / "parquet"
    output_root = tmp_path / "xlsx"
    _write_parquet(
        parquet_root / "senado__plenario_discursos.parquet",
        {
            "texto_id": ["texto-1"],
            "texto": ["Texto integral"],
            "ano": ["2026"],
        },
    )

    manifest = export_parquets_to_xlsx(
        parquet_root=parquet_root,
        output_root=output_root,
        drop_columns=["texto"],
        overwrite=True,
    )

    workbook = load_workbook(output_root / "senado__plenario_discursos.xlsx", read_only=True)
    worksheet = workbook["senado__plenario_discursos"]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("texto_id", "ano")
    assert manifest["files"][0]["columns"] == ["texto_id", "ano"]


def test_normalize_excel_value_caps_at_excel_cell_limit() -> None:
    value, stats = normalize_excel_value("x" * (EXCEL_MAX_CELL_CHARS + 1), max_cell_chars=EXCEL_MAX_CELL_CHARS)

    assert len(value) == EXCEL_MAX_CELL_CHARS
    assert stats["truncated"] == 1


def _write_parquet(path: Path, rows: dict[str, list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    pq.write_table(pa.table(rows), path)


def _args(**overrides: object) -> argparse.Namespace:
    values = {
        "profile": None,
        "parquet_root": None,
        "output_root": None,
        "parquet": None,
        "drop_column": None,
        "max_rows": None,
        "batch_size": 1000,
        "max_cell_chars": EXCEL_MAX_CELL_CHARS,
        "overwrite": False,
    }
    values.update(overrides)
    return argparse.Namespace(**values)
