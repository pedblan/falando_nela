from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from coleta.common.config import utc_now_iso
from processamento.normalizacao import DATASET_NAME, DATASET_VERSION

DEFAULT_BATCH_SIZE = 1_000
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_DATA_ROWS = EXCEL_MAX_ROWS - 1
EXCEL_MAX_CELL_CHARS = 32_767
TRUNCATION_SUFFIX = " [TRUNCADO PARA XLSX]"

FORMULA_PREFIXES = ("=", "+", "-", "@")
ILLEGAL_EXCEL_CHAR_RE = re.compile(r"[\000-\010\013\014\016-\037]")
LONG_TEXT_COLUMNS = {
    "titulo",
    "resumo",
    "indexacao",
    "texto",
    "fontes",
    "url_texto",
    "url_audio",
    "url_video",
    "url_origem",
    "raw_path",
    "raw_response_url",
}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Exporta Parquets de sample para arquivos XLSX.")
    parser.add_argument(
        "--profile",
        choices=["samples-local"],
        default=None,
        help="Preenche caminhos padrao para os Parquets locais de samples.",
    )
    parser.add_argument("--parquet-root", default=None, help="Diretorio com arquivos .parquet.")
    parser.add_argument("--output-root", default=None, help="Diretorio de saida dos arquivos .xlsx.")
    parser.add_argument(
        "--parquet",
        action="append",
        default=None,
        help="Nome ou caminho de um Parquet especifico. Pode ser repetido.",
    )
    parser.add_argument(
        "--drop-column",
        action="append",
        default=None,
        help="Coluna a remover do XLSX. Pode ser repetido, por exemplo: --drop-column texto.",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="Limita a quantidade de linhas de dados exportadas por arquivo.",
    )
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--max-cell-chars", type=int, default=EXCEL_MAX_CELL_CHARS)
    parser.add_argument("--overwrite", action="store_true")
    return parser


def main(argv: Sequence[str] | None = None) -> None:
    args = build_parser().parse_args(argv)
    parquet_root, output_root = resolve_xlsx_paths(args)
    manifest = export_parquets_to_xlsx(
        parquet_root=parquet_root,
        output_root=output_root,
        parquet_names=args.parquet,
        drop_columns=args.drop_column,
        max_rows=args.max_rows,
        batch_size=args.batch_size,
        max_cell_chars=args.max_cell_chars,
        overwrite=args.overwrite,
    )
    print(manifest["manifest_path"])


def resolve_xlsx_paths(args: argparse.Namespace) -> tuple[Path, Path]:
    if args.profile == "samples-local":
        sample_root = Path(f"data/samples/{DATASET_NAME}/{DATASET_VERSION}")
        parquet_root = Path(args.parquet_root).expanduser() if args.parquet_root else sample_root / "parquet"
        output_root = Path(args.output_root).expanduser() if args.output_root else sample_root / "xlsx"
        return parquet_root, output_root

    if not args.parquet_root:
        raise ValueError("--parquet-root e obrigatorio sem --profile")
    parquet_root = Path(args.parquet_root).expanduser()
    output_root = Path(args.output_root).expanduser() if args.output_root else parquet_root.parent / "xlsx"
    return parquet_root, output_root


def export_parquets_to_xlsx(
    *,
    parquet_root: Path,
    output_root: Path,
    parquet_names: Sequence[str] | None = None,
    drop_columns: Sequence[str] | None = None,
    max_rows: int | None = None,
    batch_size: int = DEFAULT_BATCH_SIZE,
    max_cell_chars: int = EXCEL_MAX_CELL_CHARS,
    overwrite: bool = False,
) -> dict[str, Any]:
    parquet_root = parquet_root.expanduser()
    output_root = output_root.expanduser()
    if not parquet_root.exists():
        raise FileNotFoundError(f"Diretorio de Parquets nao encontrado: {parquet_root}")
    if batch_size <= 0:
        raise ValueError("batch_size deve ser positivo")
    if max_rows is not None and max_rows <= 0:
        raise ValueError("max_rows deve ser positivo")
    if not 0 < max_cell_chars <= EXCEL_MAX_CELL_CHARS:
        raise ValueError(f"max_cell_chars deve estar entre 1 e {EXCEL_MAX_CELL_CHARS}")

    parquet_paths = list_parquet_paths(parquet_root, parquet_names)
    if not parquet_paths:
        raise FileNotFoundError(f"Nenhum arquivo .parquet encontrado em: {parquet_root}")

    manifest_path = output_root / "manifest.json"
    output_root.mkdir(parents=True, exist_ok=True)
    if manifest_path.exists() and not overwrite:
        raise FileExistsError(f"Manifest XLSX ja existe: {manifest_path}; use --overwrite para substituir.")

    drop_column_set = set(drop_columns or [])
    output_files: list[str] = []
    file_manifests: list[dict[str, Any]] = []
    total_rows = 0
    total_truncated_cells = 0
    total_sanitized_cells = 0
    total_escaped_formula_like_cells = 0

    for parquet_path in parquet_paths:
        xlsx_path = output_root / f"{parquet_path.stem}.xlsx"
        if xlsx_path.exists() and not overwrite:
            raise FileExistsError(f"Arquivo XLSX ja existe: {xlsx_path}; use --overwrite para substituir.")
        if overwrite and xlsx_path.exists():
            xlsx_path.unlink()

        result = export_parquet_to_xlsx(
            parquet_path=parquet_path,
            xlsx_path=xlsx_path,
            drop_columns=drop_column_set,
            max_rows=max_rows,
            batch_size=batch_size,
            max_cell_chars=max_cell_chars,
        )
        output_files.append(xlsx_path.as_posix())
        file_manifests.append(result)
        total_rows += result["rows_written"]
        total_truncated_cells += result["truncated_cells"]
        total_sanitized_cells += result["sanitized_cells"]
        total_escaped_formula_like_cells += result["escaped_formula_like_cells"]

    manifest = {
        "dataset": DATASET_NAME,
        "dataset_version": DATASET_VERSION,
        "created_at": utc_now_iso(),
        "parquet_root": str(parquet_root),
        "output_root": str(output_root),
        "manifest_path": str(manifest_path),
        "input_files": [path.as_posix() for path in parquet_paths],
        "input_file_count": len(parquet_paths),
        "output_files": output_files,
        "output_file_count": len(output_files),
        "rows_written": total_rows,
        "truncated_cells": total_truncated_cells,
        "sanitized_cells": total_sanitized_cells,
        "escaped_formula_like_cells": total_escaped_formula_like_cells,
        "max_rows": max_rows,
        "max_cell_chars": max_cell_chars,
        "drop_columns": sorted(drop_column_set),
        "files": file_manifests,
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return manifest


def list_parquet_paths(parquet_root: Path, parquet_names: Sequence[str] | None = None) -> list[Path]:
    if not parquet_names:
        return sorted(path for path in parquet_root.glob("*.parquet") if path.is_file())

    paths: list[Path] = []
    for name in parquet_names:
        candidate = Path(name).expanduser()
        if candidate.parent == Path(".") and not candidate.is_absolute():
            candidate = parquet_root / candidate
        if candidate.suffix != ".parquet":
            raise ValueError(f"Arquivo informado nao e .parquet: {name}")
        if not candidate.exists():
            raise FileNotFoundError(f"Parquet nao encontrado: {candidate}")
        paths.append(candidate)
    return sorted(paths)


def export_parquet_to_xlsx(
    *,
    parquet_path: Path,
    xlsx_path: Path,
    drop_columns: set[str],
    max_rows: int | None,
    batch_size: int,
    max_cell_chars: int,
) -> dict[str, Any]:
    parquet_file = pq.ParquetFile(parquet_path)
    columns = [column for column in parquet_file.schema.names if column not in drop_columns]
    if not columns:
        raise ValueError(f"Todas as colunas foram removidas para {parquet_path}")

    parquet_rows = parquet_file.metadata.num_rows
    effective_row_limit = min(max_rows, EXCEL_MAX_DATA_ROWS) if max_rows is not None else EXCEL_MAX_DATA_ROWS
    if parquet_rows > EXCEL_MAX_DATA_ROWS and max_rows is None:
        raise ValueError(
            f"{parquet_path} tem {parquet_rows} linhas; Excel suporta {EXCEL_MAX_DATA_ROWS} linhas de dados. "
            "Use --max-rows para exportar uma amostra."
        )

    workbook = Workbook(write_only=True)
    sheet_name = safe_sheet_name(parquet_path.stem)
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.freeze_panes = "A2"
    _set_column_widths(worksheet, columns)
    worksheet.append(_header_cells(worksheet, columns))

    rows_written = 0
    truncated_cells = 0
    sanitized_cells = 0
    escaped_formula_like_cells = 0

    for batch in parquet_file.iter_batches(batch_size=batch_size, columns=columns):
        for record in batch.to_pylist():
            if rows_written >= effective_row_limit:
                break
            row: list[Any] = []
            for column in columns:
                value, stats = normalize_excel_value(record.get(column), max_cell_chars=max_cell_chars)
                row.append(value)
                truncated_cells += stats["truncated"]
                sanitized_cells += stats["sanitized"]
                escaped_formula_like_cells += stats["escaped_formula_like"]
            worksheet.append(row)
            rows_written += 1
        if rows_written >= effective_row_limit:
            break

    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{rows_written + 1}"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(xlsx_path)

    return {
        "input_file": parquet_path.as_posix(),
        "output_file": xlsx_path.as_posix(),
        "sheet_name": sheet_name,
        "parquet_rows": parquet_rows,
        "rows_written": rows_written,
        "columns": columns,
        "column_count": len(columns),
        "truncated": rows_written < parquet_rows,
        "truncated_cells": truncated_cells,
        "sanitized_cells": sanitized_cells,
        "escaped_formula_like_cells": escaped_formula_like_cells,
    }


def normalize_excel_value(value: Any, *, max_cell_chars: int) -> tuple[Any, dict[str, int]]:
    stats = {"truncated": 0, "sanitized": 0, "escaped_formula_like": 0}
    if value is None or isinstance(value, (bool, int, float, datetime, date)):
        return value, stats
    if isinstance(value, Decimal):
        return str(value), stats

    if isinstance(value, bytes):
        text = value.decode("utf-8", errors="replace")
    elif isinstance(value, (dict, list, tuple)):
        text = json.dumps(value, ensure_ascii=False, sort_keys=True)
    else:
        text = str(value)

    sanitized = ILLEGAL_EXCEL_CHAR_RE.sub(" ", text)
    if sanitized != text:
        stats["sanitized"] = 1
    if sanitized.startswith(FORMULA_PREFIXES):
        sanitized = "'" + sanitized
        stats["escaped_formula_like"] = 1
    if len(sanitized) > max_cell_chars:
        suffix = TRUNCATION_SUFFIX if max_cell_chars > len(TRUNCATION_SUFFIX) else ""
        sanitized = sanitized[: max_cell_chars - len(suffix)] + suffix
        stats["truncated"] = 1
    return sanitized, stats


def safe_sheet_name(value: str) -> str:
    sheet_name = re.sub(r"[][*/\\?:]", "_", value).strip("'")
    sheet_name = sheet_name or "dados"
    return sheet_name[:31]


def _header_cells(worksheet: Any, columns: Sequence[str]) -> list[WriteOnlyCell]:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    font = Font(bold=True)
    cells: list[WriteOnlyCell] = []
    for column in columns:
        cell = WriteOnlyCell(worksheet, value=column)
        cell.font = font
        cell.fill = fill
        cells.append(cell)
    return cells


def _set_column_widths(worksheet: Any, columns: Sequence[str]) -> None:
    for index, column in enumerate(columns, start=1):
        width = 70 if column in LONG_TEXT_COLUMNS else min(max(len(column) + 2, 12), 30)
        worksheet.column_dimensions[get_column_letter(index)].width = width


if __name__ == "__main__":
    main()
